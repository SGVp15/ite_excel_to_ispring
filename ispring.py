import os
from copy import deepcopy

import openpyxl

from Question import Question
from TICKET.ticket import Ticket
from config import template_file_for_ispring, OUTPUT_DIR
from list_utils import compress


def create_txt_file_category(ticket: Ticket, file_path: str, max_questions_in_ticket=30):
    questions = ticket.questions
    questions_by_category = ticket.questions_by_category

    with open(file_path, 'w', encoding='utf-8') as f:
        count_all_sum = 0
        categories = sorted(questions_by_category.keys())

        list_max = [len(questions_by_category.get(category)) for category in categories]
        list_a = deepcopy(list_max)
        list_a, list_max = compress(list_a, list_max, max_questions_in_ticket)
        for i, category in enumerate(categories):
            f.write(f'{category}\t{list_a[i]}\t{list_max[i]}\n')
        if sum(list_max) < max_questions_in_ticket:
            f.write(f'\n{count_all_sum} Всего вопросов: {len(questions)}')


def create_excel_file_for_ispring(ticket: Ticket, exam_name: str, path_out: str, num_box=0):
    head = read_template()
    for category, question_list in ticket.questions_by_category.items():
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for i, v in enumerate(head):
            worksheet.cell(row=1, column=i + 1, value=str(v))

        row = 0
        # Create a list of values to write to the Excel file
        for q in question_list:
            row += 1
            values_to_write = ['MC', q.text_question, os.path.basename(q.image), '', '',
                               f'*{q.ans_a}', q.ans_b, q.ans_c, q.ans_d,
                               '', '', '', '', '', '', '', '', 1]

            # Write the numbers to the worksheet
            for i, v in enumerate(values_to_write):
                worksheet.cell(row=row + 1, column=i + 1, value=str(v))

        workbook.save(os.path.join(path_out, f'{category[:2]}.xlsx'))


def create_tickets(questions: [Question]) -> [Ticket]:
    tickets = []
    ticket = Ticket(questions)
    max_box = ticket.get_max_len_box()
    for i in range(max_box):
        tickets.append(ticket.create_unique_ticket(i))
    return tickets


def read_template(file=template_file_for_ispring):
    if os.path.exists(file):
        workbook = openpyxl.load_workbook(file)
        page_name = workbook.sheetnames
        worksheet = workbook[page_name[0]]
        for row in worksheet.iter_rows(values_only=True):
            return row
