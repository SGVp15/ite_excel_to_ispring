import json
import os
import random
import re

import openpyxl
from openpyxl.utils import get_column_letter

from Question import Question


def get_all_questions_from_excel_file(file: str) -> [Question]:
    wb = openpyxl.load_workbook(filename=f'{file}', data_only=True)
    page_name = wb.sheetnames
    page_name = str(page_name[0])

    map_excel = {}
    max_column = wb[page_name].max_column
    for col in range(1, max_column):
        map_excel[read_excel(wb, page_name, get_column_letter(col), 1)] = get_column_letter(col)
    column_id_question = map_excel['Код вопроса']
    column_category_question = map_excel['Раздел курса']
    column_box_question = map_excel['Блок вопросов']
    column_enable_question = map_excel.get('Действующий 1-да, 0-нет', 'L')

    column_a = 'a'
    column_main = 'b'

    all_questions = []
    num_q = 0

    i = 0
    max_row = wb[page_name].max_row
    while i <= max_row:
        i += 1
        a = read_excel(wb, page_name, column_a, i + 1).lower()
        if a not in ('a', 'а'):
            continue

        b = read_excel(wb, page_name, column_a, i + 2).lower()
        c = read_excel(wb, page_name, column_a, i + 3).lower()
        d = read_excel(wb, page_name, column_a, i + 4).lower()

        if b not in ('b', 'в') or c not in ('c', 'с') or d != 'd':
            continue

        enable_question = read_excel(wb, page_name, column_enable_question, i)
        if enable_question in ('1', 1):
            q = Question()

            q.id_question = read_excel(wb, page_name, column_id_question, i)
            q.box_question = read_excel(wb, page_name, column_box_question, i)

            image_pattern = r'\s*\[(.*)\]\s*'
            if re.search(image_pattern, q.box_question):
                q.image = re.findall(pattern=image_pattern, string=q.box_question)[0]
                q.image = q.image.strip()
            else:
                q.image = ''

            if q.box_question == 'None':
                q.box_question = None

            q.text_question = read_excel(wb, page_name, column_main, i)
            q.answer_a = read_excel(wb, page_name, column_main, i + 1)
            q.answer_b = read_excel(wb, page_name, column_main, i + 2)
            q.answer_c = read_excel(wb, page_name, column_main, i + 3)
            q.answer_d = read_excel(wb, page_name, column_main, i + 4)
            num_q += 1
            q.category = read_excel(wb, page_name, column_category_question, i)
            all_questions.append(q)
        i += 3

    return all_questions


def save_json_questions(path_questions, all_questions):
    file_json = os.path.join(path_questions, f'{all_questions[0].exam}.json')
    with open(file_json, 'w', encoding='utf-8') as f:
        f.write(json.dumps(all_questions))


def read_excel(excel, page_name, column, row):
    sheet_ranges = excel[page_name]
    value = str(sheet_ranges[f'{column}{row}'].value).strip()
    return value
